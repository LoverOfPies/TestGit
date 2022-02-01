import openpyxl


def get_data():
    wb = openpyxl.load_workbook('./gr.xlsx')
    sheet = wb["33"]
    data = []
    result = []
    for i in range(2, sheet.max_row):
        data.append(sheet.cell(row=i, column=1).value)
    for i in range(0, len(data)):
        resstr = data[i]
        for j in range(i + 1, len(data)):
            if (data[i] in data[j]):
                resstr = resstr + " -" + data[j].replace(data[i], "")
        result.append(resstr)
    for i in range(1, len(result)):
        sheet.cell(row=i, column=2).value = result[i-1]
    wb.save('./gr.xlsx')
    print(result)


if __name__ == '__main__':
    get_data()
